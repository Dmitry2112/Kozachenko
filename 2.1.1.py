import csv
import re
import openpyxl
import openpyxl.utils
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font


class Report:
    def __init__(self, job_name):
        self.wb = openpyxl.Workbook()
        self.ws1 = self.wb.create_sheet('Статистика по годам', 0)
        self.ws2 = self.wb.create_sheet('Статистика по городам', 1)
        custom_font = Font(bold=True)
        self.cell_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                  bottom=Side(style='thin'))
        self.ws1['A1'], self.ws1['A1'].font, self.ws1['A1'].border = 'Год', custom_font, self.cell_border
        self.ws1['B1'], self.ws1['B1'].font, self.ws1['B1'].border = 'Средняя зарплата', custom_font, self.cell_border
        self.ws1['C1'], self.ws1['C1'].font, self.ws1['C1'].border = f'Средняя зарплата - {job_name}', custom_font, self.cell_border
        self.ws1['D1'], self.ws1['D1'].font, self.ws1['D1'].border = 'Количество вакансий', custom_font, self.cell_border
        self.ws1['E1'], self.ws1['E1'].font, self.ws1['E1'].border = f'Количество вакансий - {job_name}', custom_font, self.cell_border
        self.ws2['A1'], self.ws2['A1'].font, self.ws2['A1'].border = 'Город', custom_font, self.cell_border
        self.ws2['B1'], self.ws2['B1'].font, self.ws2['B1'].border = 'Уровень зарплат', custom_font, self.cell_border
        self.ws2['D1'], self.ws2['D1'].font, self.ws2['D1'].border = 'Город', custom_font, self.cell_border
        self.ws2['E1'], self.ws2['E1'].font, self.ws2['E1'].border = 'Доля вакансий', custom_font, self.cell_border


    def generate_excel(self, all_dict):
        row_index = 2
        for y in all_dict[0]:
            self.ws1.cell(column=1, row=row_index, value=y)
            self.ws1.cell(column=1, row=row_index).border = self.cell_border
            for i in range(0, 4):
                self.ws1.cell(column=2+i, row=row_index, value=all_dict[i][y])
                self.ws1.cell(column=2+i, row=row_index).border = self.cell_border
            row_index += 1
        row_index = 2
        for city in all_dict[4]:
            self.ws2.cell(column=1, row=row_index, value=city)
            self.ws2.cell(column=1, row=row_index).border = self.cell_border
            self.ws2.cell(column=2, row=row_index, value=all_dict[4][city])
            self.ws2.cell(column=2, row=row_index).border = self.cell_border
            row_index += 1
            if row_index == 12:
                break
        row_index = 2
        for city in all_dict[5]:
            self.ws2.cell(column=4, row=row_index, value=city)
            self.ws2.cell(column=4, row=row_index).border = self.cell_border
            self.ws2.cell(column=5, row=row_index, value=all_dict[5][city])
            self.ws2.cell(column=5, row=row_index).number_format = '0.00%'
            self.ws2.cell(column=5, row=row_index).border = self.cell_border
            row_index += 1
            if row_index == 12:
                break

        for column_cells in self.ws1.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = length * 1.2  #Коэффициент, чтобы избавиться от проблем связанных с разными шрифтами
            self.ws1.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        self.ws2.column_dimensions['C'].width = 5
        for column_cells in self.ws2.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = length * 1.2
            self.ws2.column_dimensions[column_cells[0].column_letter].width = adjusted_width
        self.wb.save('report.xlsx')


def clean_string(string):
    string = re.sub(r'<[^>]*>', '', string)
    string = ' '.join(string.split())
    return string


def collect_data(file_name, job_name):
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    file_csv = csv.reader(open(file_name, encoding='utf_8_sig'))
    list_data = [x for x in file_csv]
    titles = list_data[0]
    values = [x for x in list_data[1:] if x.count('') == 0 and len(x) == len(titles)]
    total_vac_per_year = {}
    salary_per_year = {}
    job_vac_per_year = {}
    job_salary_per_year = {}
    salary_per_city = {}
    vac_per_city = {}
    limit_vac_num = int(len(values) * 0.01)
    name_index = 0
    salary_from_index = 0
    salary_to_index = 0
    salary_currency_index = 0
    area_name_index = 0
    published_at_index = 0
    for i in range(len(titles)):
        if titles[i] == 'name':
            name_index = i
        if titles[i] == 'salary_from':
            salary_from_index = i
        if titles[i] == 'salary_to':
            salary_to_index = i
        if titles[i] == 'salary_currency':
            salary_currency_index = i
        if titles[i] == 'area_name':
            area_name_index = i
        if titles[i] == 'published_at':
            published_at_index = i

    for k in range(len(values)):
        raw_vac = values[k]
        vac = []
        for j in range(len(raw_vac)):
            vac.append(clean_string(raw_vac[j]))
        year = int(vac[published_at_index][0:4])
        salary_rub = (float(vac[salary_from_index]) + float(vac[salary_to_index])) * 0.5 * currency_to_rub[vac[salary_currency_index]]
        if year in total_vac_per_year:
            total_vac_per_year[year] += 1
        else:
            total_vac_per_year[year] = 1
        #job_name
        if job_name in vac[name_index]:
            if year in job_vac_per_year:
                job_vac_per_year[year] += 1
            else:
                job_vac_per_year[year] = 1
            if year in job_salary_per_year:
                x = job_salary_per_year[year]
                x = x[0] + salary_rub, x[1] + 1
                job_salary_per_year[year] = x
            else:
                job_salary_per_year[year] = salary_rub, 1

        if year in salary_per_year:
            x = salary_per_year[year]
            x = x[0] + salary_rub, x[1] + 1
            salary_per_year[year] = x
        else:
            salary_per_year[year] = salary_rub, 1

        #per_city
        if vac[area_name_index] in vac_per_city:
            vac_per_city[vac[area_name_index]] += 1
        else:
            vac_per_city[vac[area_name_index]] = 1

        if vac[area_name_index] in salary_per_city:
            x = salary_per_city[vac[area_name_index]]
            x = x[0] + salary_rub, x[1] + 1
            salary_per_city[vac[area_name_index]] = x
        else:
            salary_per_city[vac[area_name_index]] = salary_rub, 1

    for y in salary_per_year:
        x = salary_per_year[y]
        salary_per_year[y] = int(x[0] / x[1])

    for y in job_salary_per_year:
        x = job_salary_per_year[y]
        job_salary_per_year[y] = int(x[0] / x[1])

    for y in salary_per_city:
        x = salary_per_city[y]
        salary_per_city[y] = int(x[0] / x[1])

    for y in salary_per_year:
        if y not in job_salary_per_year:
            job_salary_per_year[y] = 0
    
    for y in salary_per_year:
        if y not in job_vac_per_year:
            job_vac_per_year[y] = 0

    filtered_salary_per_city = {}
    for s in salary_per_city:
        if vac_per_city[s] >= limit_vac_num:
            filtered_salary_per_city[s] = salary_per_city[s]

    filtered_vac_per_city = {}
    for s in vac_per_city:
        if vac_per_city[s] >= limit_vac_num:
            filtered_vac_per_city[s] = vac_per_city[s]

    final_salary_per_city = dict(sorted(filtered_salary_per_city.items(), key=lambda x: x[1], reverse=True))
    final_vac_per_city = dict(sorted(filtered_vac_per_city.items(), key=lambda x: x[1], reverse=True))

    for city in final_vac_per_city:
        final_vac_per_city[city] /= len(values)

    return [salary_per_year, job_salary_per_year, total_vac_per_year, job_vac_per_year, final_salary_per_city, final_vac_per_city]


file_name = input('Введите название файла: ')
job_name = input('Введите название профессии: ')
rep = Report(job_name)
all_dict = collect_data(file_name, job_name)
rep.generate_excel(all_dict)